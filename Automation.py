import openpyxl
import os

# Check if the data.xlsx file exists
if os.path.isfile('data.xlsx'):
    workbook = openpyxl.load_workbook('data.xlsx')
else:
    workbook = openpyxl.Workbook()

# Select the desired sheet or create a new one
sheet = workbook['Sheet1'] if 'Sheet1' in workbook.sheetnames else workbook.create_sheet('Sheet1')

# Column names
column_names = ['Name', 'Email', 'Phone']

# Enter column names in the first row if it's a new file
if not sheet['A1'].value:
    for idx, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=idx).value = column_name

# Sample data to enter
data = [
    {'row': 2, 'column': 1, 'value': 'John'},
    {'row': 2, 'column': 2, 'value': 'johndoe@example.com'},
    {'row': 2, 'column': 3, 'value': '1234567890'},
    # Add more data entries as needed
]

# Generate sample data for 100 more people
for i in range(100):
    row = 3 + i
    name = f'Person {i+1}'
    email = f'person{i+1}@example.com'
    phone = f'123456789{i+1}'

    data.append({'row': row, 'column': 1, 'value': name})
    data.append({'row': row, 'column': 2, 'value': email})
    data.append({'row': row, 'column': 3, 'value': phone})

# Loop through the data and enter it
for entry in data:
    row = entry['row']
    column = entry['column']
    value = entry['value']

    # Set the cell value
    sheet.cell(row=row, column=column).value = value

# Save the changes to the workbook
workbook.save('data.xlsx')
